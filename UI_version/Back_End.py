from openpyxl import Workbook, load_workbook

grades = "Sample Grades DB.xlsx"
wb1 = load_workbook(grades)
ws1 = wb1.active

class faculty:
    def __init__(self,name,id):
        self.name = name
        self.id = id
    
class student:
    def __init__(self,name,id):
        self.name = name
        self.id = id
    


def print_grades(student_id):
    for row in ws1.iter_rows(min_row=2,values_only=True):
        if row[1] == student_id:
            return {"sub1":row[2],"sub2":row[3],"sub3":row[4],"sub4":row[5],"sub5":row[6],"sub6":row[7],"sub7":row[8],"sub8":row[9]}
    return False

def print_all_students():
    Data = {}
    for row in ws1.iter_rows(min_row=2,values_only=True):
        if row[0] != None:
            Data[row[0]] = [row[1],row[2],row[3],row[4],row[5],row[6],row[7],row[8],row[9]]
        else:
            continue
    return Data